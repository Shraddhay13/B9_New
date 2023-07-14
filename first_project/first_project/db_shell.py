from app1.models import Student

# exec(open(r"G:\Python-B9\Files\DjangoProjects\first_project\first_project\db_shell.py").read())

# model manager -- methods :- all(), bulk_create(), get(), filter()

# all_studs = Student.objects.all()[0:3]  # returns a queryset --list -- select * from student;
# all_studs = Student.objects.all()
# print(all_studs.__dict__)
# print(all_studs)
# print(all_studs.query)  # To know the sql query

# for stud in all_studs:
#     print(stud.name)

# for stud in all_studs:
#     print(f"--------Stud ID:- {stud.id}-------------")
#     print(stud.show_detail())
    # print("-------------------------")


# print(list(all_studs))

# stud_obj = Student.objects.get(address="Pune")  # always return single record
# print(stud_obj)
# print(stud_obj.show_detail())

# update student set name="CC", age=35 where id=8;

# stud = Student.objects.get(id=8)
# stud.name = "CC"
# stud.age = 35
# stud.save()

# stud = Student.objects.get(id=8)
# stud.delete()

# print(dir(Student.objects))  - To check object methods

# stud = Student.objects.values()   # queryset of objects dict
# print(stud)
# for i in stud:
#     print(i)

# studs = Student.objects.values_list()
# for i in studs:
#     print(i)


# Student.objects.get(id=9).delete()


# try:
#     stud_obj = Student.objects.get(address="Pune", name="A")  
# except Student.MultipleObjectsReturned:
#     print("More that one record found")
# except Student.DoesNotExist:
#     print("Student does not exist with given arguments")



# stud = Student.objects.filter(address="Pune")  # returns queryset -- if data is not available then it will return empty queryset
# print(stud)


# 2nd way to create object
# s2 = Student(name="Payal", marks=35, address="Aurangabad", age=29)
# s2.save()

# 1st way to create object 
# Student.objects.create(name="Sonali", marks=68, address="Banglore", age=27)

# s1 =Student(name="A", marks=99, address="Adr1", age=21)
# s2 =Student(name="B", marks=69, address="Adr2", age=19)
# s3 =Student(name="C", marks=19, address="Adr3", age=31)

# Student.objects.bulk_create([s1, s2, s3])


# data = Student.objects.first()
# print(data)

# print(Student.objects.count())

# SQLite3 , Mysql - Workbench , PostgreSQL -pgAdmin4 - using psycopg2
# - using python , django using PostgreSQL


# records = Student.objects.filter(name__startswith="S")
# print(records)


# records = Student.objects.filter(name__in=["Shraddha", "A", "B"])
# print(records)

# print(len(Student.objects.exclude(address="Pune")))

# 1st way for OR operator
# queryset = Student.objects.filter(name__startswith= "P") | Student.objects.filter(address__startswith="A")
# print(queryset)

from django.db.models import Q, F

# 2nd way for OR operator
# queryset = Student.objects.filter(Q(name__startswith= "P") | Q(address__startswith="A"))
# print(queryset)

# queryset = Student.objects.filter(name="P" , address ="A")
# print(queryset)


# data = Student.objects.filter(Q(name__startswith='A')| (Q(address='Mumbai')&Q(marks=69)))
# print(data)

# data = Student.objects.filter(Q(age__gt=20) & Q(age__lt=40))
# print(data)

# Student.objects.all().update(marks=F('marks')+5)

# print(Student.objects.filter(name__contains="P" , address ="A")t)

# Student.objects.filter(name="A").update(name="AA")

# print(Student.objects.all().order_by('-marks'))  # ASC & DESCC order

from django.db.models import Avg, Max, Min, Sum, Count

# print(round(Student.objects.all().aggregate(Avg('age')).get("age__avg")))

# q1 = Student.objects.filter(id__lt = 4)
# q2 = Student.objects.filter(age__gt = 40)

# result = q1.union(q2)
# print(result)


from openpyxl import Workbook, load_workbook, worksheet

# wb = load_workbook(r"G:\Python-B9\Files\DjangoProjects\first_project\Stud.xlsx")
# sheet = wb.active

# # print(sheet.max_row)

# lst =[]
# for i in range(2, sheet.max_row + 1):
#     nm = sheet.cell(row=i, column=1).value
#     mrk = sheet.cell(row=i, column=2).value
#     adr = sheet.cell(row=i, column=3).value
#     ag = sheet.cell(row=i, column=4).value

#     obj=Student(name=nm, marks=mrk, address=adr, age=ag)
#     lst.append(obj)

# Student.objects.bulk_create(lst)

# To run the raw sql queries using django

# data = Student.objects.raw("select * from student")
# for i in data:
#     print(i)


from django.db import connection

# print(connection)
# cursor = connection.cursor()
# cursor.execute("select * from student where name='AA'")
# data = cursor.fetchall()
# print(data)